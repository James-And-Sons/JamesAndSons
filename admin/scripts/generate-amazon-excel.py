import sys
import os
import json
import re
import openpyxl

def extract_number(text):
    if not text:
        return None
    # Find the first float or int in the string
    match = re.search(r"[-+]?\d*\.\d+|\d+", str(text))
    if match:
        try:
            return float(match.group())
        except ValueError:
            return None
    return None

def main():
    if len(sys.argv) < 3:
        print("Usage: python generate-amazon-excel.py <input_json_path> <output_xlsm_path>")
        sys.exit(1)

    input_json_path = sys.argv[1]
    output_xlsm_path = sys.argv[2]

    # Load input data
    with open(input_json_path, "r", encoding="utf-8") as f:
        products = json.load(f)

    # Load template xlsm
    template_path = "/Users/abhishikt_mac/Skills/Coding/Growth-ho clients/JamesAndSons/LAMP_LIGHT_FIXTURE.xlsm"
    if not os.path.exists(template_path):
        print(f"Error: Template file not found at {template_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(template_path, keep_vba=True)
    if "Template" not in wb.sheetnames:
        print("Error: 'Template' sheet not found in the workbook")
        sys.exit(1)

    sheet = wb["Template"]

    # Clear any old data starting from Row 8
    # Keep the structure intact but clear cells from Row 8 to the max row
    max_row = sheet.max_row
    if max_row >= 8:
        for r in range(8, max_row + 1):
            for c in range(1, sheet.max_column + 1):
                sheet.cell(r, c).value = None

    row_idx = 8

    for p in products:
        variants = p.get("variants", [])
        has_variants = len(variants) > 0

        # Helper to get value from variant or parent fallback
        def get_val(var, key, fallback=None):
            if var and key in var and var[key] is not None:
                return var[key]
            return p.get(key, fallback)

        # Write Parent row if variants exist
        if has_variants:
            p_brand = p.get("brand") or "James and Sons"
            p_desc = p.get("description") or p.get("name")
            p_material = p.get("material") or (p.get("materialAndFinish")[0] if p.get("materialAndFinish") else None)
            p_origin = p.get("countryOfOrigin") or "India"

            # Write parent details
            sheet.cell(row_idx, 1).value = p.get("sku") # SKU
            sheet.cell(row_idx, 2).value = "LIGHT_FIXTURE" # Product Type
            sheet.cell(row_idx, 3).value = "Create or Replace (Full Update)" # Listing Action
            sheet.cell(row_idx, 4).value = "Parent" # Parentage Level
            sheet.cell(row_idx, 6).value = "Size/Color" # Variation Theme Name
            sheet.cell(row_idx, 7).value = p.get("name") # Item Name
            sheet.cell(row_idx, 9).value = p_brand # Brand Name
            # Image URL fallbacks
            p_images = p.get("images") or []
            if len(p_images) > 0:
                sheet.cell(row_idx, 22).value = p_images[0]
                for idx, img in enumerate(p_images[1:9]):
                    sheet.cell(row_idx, 23 + idx).value = img
            sheet.cell(row_idx, 32).value = p_desc # Product Description
            # Bullets
            p_bullets = p.get("bulletPoints") or []
            for idx, b in enumerate(p_bullets[:5]):
                sheet.cell(row_idx, 33 + idx).value = b
            if p_material:
                sheet.cell(row_idx, 49).value = p_material # Material
            # Wattage
            p_watt = extract_number(p.get("power"))
            if p_watt is not None:
                sheet.cell(row_idx, 83).value = p_watt
                sheet.cell(row_idx, 84).value = "Watts"
            # Voltage
            p_volt = extract_number(p.get("voltage"))
            if p_volt is not None:
                sheet.cell(row_idx, 85).value = p_volt
                sheet.cell(row_idx, 86).value = "Volts"
            sheet.cell(row_idx, 312).value = p_origin # Country of Origin
            
            row_idx += 1

            # Write Children rows
            for v in variants:
                v_sku = v.get("sku")
                v_price = get_val(v, "d2cPrice")
                v_mrp = get_val(v, "mrp")
                v_stock = get_val(v, "stockQuantity") or 0
                v_images = v.get("images") or p.get("images") or []
                v_weight = get_val(v, "weight") or 0.5
                v_length = get_val(v, "length") or 10.0
                v_width = get_val(v, "breadth") or 10.0
                v_height = get_val(v, "height") or 10.0
                v_watt = extract_number(get_val(v, "power"))
                v_volt = extract_number(get_val(v, "voltage"))
                v_material = get_val(v, "material") or (p.get("materialAndFinish")[0] if p.get("materialAndFinish") else None)
                v_origin = get_val(v, "countryOfOrigin") or "India"
                v_brand = get_val(v, "brand") or "James and Sons"
                v_bullets = v.get("bulletPoints") or p.get("bulletPoints") or []

                sheet.cell(row_idx, 1).value = v_sku # SKU
                sheet.cell(row_idx, 2).value = "LIGHT_FIXTURE" # Product Type
                sheet.cell(row_idx, 3).value = "Create or Replace (Full Update)" # Listing Action
                sheet.cell(row_idx, 4).value = "Child" # Parentage Level
                sheet.cell(row_idx, 5).value = p.get("sku") # Parent SKU
                sheet.cell(row_idx, 6).value = "Size/Color" # Variation Theme Name
                sheet.cell(row_idx, 7).value = f"{p.get('name')} - {v.get('name')}" # Item Name
                sheet.cell(row_idx, 9).value = v_brand # Brand Name
                sheet.cell(row_idx, 10).value = "SellerSKU" # Product Id Type
                sheet.cell(row_idx, 11).value = v_sku # Product Id
                # Images
                if len(v_images) > 0:
                    sheet.cell(row_idx, 22).value = v_images[0]
                    for idx, img in enumerate(v_images[1:9]):
                        sheet.cell(row_idx, 23 + idx).value = img
                sheet.cell(row_idx, 32).value = p_desc # Product Description
                # Bullets
                for idx, b in enumerate(v_bullets[:5]):
                    sheet.cell(row_idx, 33 + idx).value = b
                if v_material:
                    sheet.cell(row_idx, 49).value = v_material # Material
                if v_watt is not None:
                    sheet.cell(row_idx, 83).value = v_watt
                    sheet.cell(row_idx, 84).value = "Watts"
                if v_volt is not None:
                    sheet.cell(row_idx, 85).value = v_volt
                    sheet.cell(row_idx, 86).value = "Volts"
                # Weight
                sheet.cell(row_idx, 197).value = v_weight
                sheet.cell(row_idx, 198).value = "kg"
                # Dimensions
                sheet.cell(row_idx, 229).value = v_height
                sheet.cell(row_idx, 230).value = "cm"
                sheet.cell(row_idx, 231).value = v_length
                sheet.cell(row_idx, 232).value = "cm"
                sheet.cell(row_idx, 233).value = v_width
                sheet.cell(row_idx, 234).value = "cm"
                
                sheet.cell(row_idx, 269).value = v_stock # Quantity (IN)
                sheet.cell(row_idx, 273).value = v_price # Your Price INR
                sheet.cell(row_idx, 274).value = v_mrp # Maximum Retail Price
                sheet.cell(row_idx, 312).value = v_origin # Country of Origin

                row_idx += 1
        else:
            # Single product with no variants
            p_sku = p.get("sku")
            p_price = p.get("d2cPrice")
            p_mrp = p.get("mrp")
            p_stock = p.get("stockQuantity") or 0
            p_images = p.get("images") or []
            p_brand = p.get("brand") or "James and Sons"
            p_desc = p.get("description") or p.get("name")
            p_bullets = p.get("bulletPoints") or []
            p_material = p.get("material") or (p.get("materialAndFinish")[0] if p.get("materialAndFinish") else None)
            p_watt = extract_number(p.get("power"))
            p_volt = extract_number(p.get("voltage"))
            p_weight = p.get("weight") or 0.5
            p_length = p.get("length") or 10.0
            p_width = p.get("breadth") or 10.0
            p_height = p.get("height") or 10.0
            p_origin = p.get("countryOfOrigin") or "India"

            sheet.cell(row_idx, 1).value = p_sku # SKU
            sheet.cell(row_idx, 2).value = "LIGHT_FIXTURE" # Product Type
            sheet.cell(row_idx, 3).value = "Create or Replace (Full Update)" # Listing Action
            sheet.cell(row_idx, 4).value = None # Parentage Level
            sheet.cell(row_idx, 7).value = p.get("name") # Item Name
            sheet.cell(row_idx, 9).value = p_brand # Brand Name
            sheet.cell(row_idx, 10).value = "SellerSKU" # Product Id Type
            sheet.cell(row_idx, 11).value = p_sku # Product Id
            # Images
            if len(p_images) > 0:
                sheet.cell(row_idx, 22).value = p_images[0]
                for idx, img in enumerate(p_images[1:9]):
                    sheet.cell(row_idx, 23 + idx).value = img
            sheet.cell(row_idx, 32).value = p_desc # Product Description
            # Bullets
            for idx, b in enumerate(p_bullets[:5]):
                sheet.cell(row_idx, 33 + idx).value = b
            if p_material:
                sheet.cell(row_idx, 49).value = p_material # Material
            if p_watt is not None:
                sheet.cell(row_idx, 83).value = p_watt
                sheet.cell(row_idx, 84).value = "Watts"
            if p_volt is not None:
                sheet.cell(row_idx, 85).value = p_volt
                sheet.cell(row_idx, 86).value = "Volts"
            # Weight
            sheet.cell(row_idx, 197).value = p_weight
            sheet.cell(row_idx, 198).value = "kg"
            # Dimensions
            sheet.cell(row_idx, 229).value = p_height
            sheet.cell(row_idx, 230).value = "cm"
            sheet.cell(row_idx, 231).value = p_length
            sheet.cell(row_idx, 232).value = "cm"
            sheet.cell(row_idx, 233).value = p_width
            sheet.cell(row_idx, 234).value = "cm"

            sheet.cell(row_idx, 269).value = p_stock # Quantity (IN)
            sheet.cell(row_idx, 273).value = p_price # Your Price INR
            sheet.cell(row_idx, 274).value = p_mrp # Maximum Retail Price
            sheet.cell(row_idx, 312).value = p_origin # Country of Origin

            row_idx += 1

    wb.save(output_xlsm_path)
    print(f"Successfully generated populated Excel sheet at: {output_xlsm_path}")

if __name__ == "__main__":
    main()
